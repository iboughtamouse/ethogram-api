#!/usr/bin/env python3
"""
Detailed Excel formatting analysis - check every cell in both sheets
"""

import openpyxl
from openpyxl.utils import get_column_letter

def detailed_analysis(original_file, generated_file):
    """Detailed cell-by-cell comparison"""
    
    orig_wb = openpyxl.load_workbook(original_file)
    gen_wb = openpyxl.load_workbook(generated_file)
    
    orig_ws = orig_wb[orig_wb.sheetnames[0]]
    gen_ws = gen_wb[gen_wb.sheetnames[0]]
    
    print("=" * 80)
    print("DETAILED FORMATTING ANALYSIS")
    print("=" * 80)
    print()
    
    # Frozen panes
    print("🧊 FROZEN PANES")
    print("-" * 80)
    print(f"Original:  {orig_ws.freeze_panes}")
    print(f"Generated: {gen_ws.freeze_panes}")
    print(f"Expected:  Freeze at cell B5 (freeze row 5 and column A)")
    print()
    
    # Column widths
    print("📏 COLUMN WIDTHS")
    print("-" * 80)
    for col in range(1, 14):
        col_letter = get_column_letter(col)
        orig_width = orig_ws.column_dimensions[col_letter].width or 8.43  # Default Excel width
        gen_width = gen_ws.column_dimensions[col_letter].width or 8.43
        status = "✅" if abs(orig_width - gen_width) < 0.1 else "❌"
        print(f"Column {col_letter:>2}: Original={orig_width:>6.2f}  Generated={gen_width:>6.2f}  {status}")
    print()
    
    # Header rows (rows 1-4)
    print("📋 HEADER FORMATTING (Rows 1-4)")
    print("-" * 80)
    for row in range(1, 5):
        for col in range(1, 14):
            orig_cell = orig_ws.cell(row, col)
            gen_cell = gen_ws.cell(row, col)
            
            if orig_cell.value:
                col_letter = get_column_letter(col)
                cell_ref = f"{col_letter}{row}"
                
                # Compare formatting
                issues = []
                
                if orig_cell.font and orig_cell.font.bold != (gen_cell.font.bold if gen_cell.font else None):
                    issues.append(f"Bold: {orig_cell.font.bold} vs {gen_cell.font.bold if gen_cell.font else None}")
                
                if orig_cell.alignment and orig_cell.alignment.wrap_text != (gen_cell.alignment.wrap_text if gen_cell.alignment else None):
                    issues.append(f"Wrap: {orig_cell.alignment.wrap_text} vs {gen_cell.alignment.wrap_text if gen_cell.alignment else None}")
                
                if orig_cell.alignment and orig_cell.alignment.horizontal != (gen_cell.alignment.horizontal if gen_cell.alignment else None):
                    issues.append(f"Align: {orig_cell.alignment.horizontal} vs {gen_cell.alignment.horizontal if gen_cell.alignment else None}")
                
                if issues:
                    print(f"  {cell_ref}: {orig_cell.value[:40]}")
                    for issue in issues:
                        print(f"    ❌ {issue}")
    print()
    
    # Time slot rows (row 4 is header, data starts at row 5)
    print("⏰ TIME SLOT FORMATTING (Data rows)")
    print("-" * 80)
    print("Checking first 3 data rows for formatting patterns...")
    for row in range(5, 8):
        orig_time = orig_ws.cell(row, 2).value  # Column B is time
        gen_time = gen_ws.cell(row, 2).value
        
        print(f"  Row {row}: {orig_time} | {gen_time}")
        
        for col in range(1, 14):
            orig_cell = orig_ws.cell(row, col)
            gen_cell = gen_ws.cell(row, col)
            
            # Check wrap text on data cells
            if orig_cell.value and orig_cell.alignment and orig_cell.alignment.wrap_text:
                gen_wrap = gen_cell.alignment.wrap_text if gen_cell.alignment else None
                if not gen_wrap:
                    col_letter = get_column_letter(col)
                    print(f"    ❌ {col_letter}: Original has wrap_text=True, Generated has {gen_wrap}")
    print()
    
    # Summary
    print("=" * 80)
    print("SUMMARY OF REQUIRED FIXES")
    print("=" * 80)
    print()
    print("1. ❌ FROZEN PANES")
    print("   Need: worksheet.freeze_panes = 'B5'")
    print("   This freezes the top 4 header rows and column A")
    print()
    print("2. ❌ COLUMN WIDTHS")
    print("   Need:")
    print("   - Column A (Date/Time): 25.75")
    print("   - Column B (Time): 4.88")
    print("   - Columns C-M (Data): 13.0 each")
    print()
    print("3. ❌ HEADER ROW FORMATTING (Rows 1-4)")
    print("   Need:")
    print("   - Row 1, A1: Bold title")
    print("   - Row 3, B3: Bold 'Time:' label")
    print("   - Row 4: Bold headers")
    print()
    print("4. ❓ TEXT WRAPPING (Check data cells)")
    print("   May need wrap_text=True on data cells")
    print()

if __name__ == '__main__':
    original = '1 Hour Ethogram Section(2).xlsx'
    generated = 'test-formatted-output.xlsx'  # Our newly formatted version
    
    try:
        detailed_analysis(original, generated)
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
