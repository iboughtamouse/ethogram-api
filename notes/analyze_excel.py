#!/usr/bin/env python3
"""
Analyze Excel formatting differences between original and generated spreadsheets
"""

import openpyxl
from openpyxl.utils import get_column_letter
import json

def analyze_workbook(filename):
    """Extract formatting information from an Excel workbook"""
    wb = openpyxl.load_workbook(filename)
    analysis = {
        'filename': filename,
        'sheets': {}
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            'dimensions': {
                'max_row': ws.max_row,
                'max_column': ws.max_column,
            },
            'frozen_panes': str(ws.freeze_panes) if ws.freeze_panes else None,
            'column_widths': {},
            'row_heights': {},
            'merged_cells': [str(cell) for cell in ws.merged_cells],
            'sample_cells': {}
        }
        
        # Get column widths
        for col in range(1, min(ws.max_column + 1, 20)):  # First 20 columns
            col_letter = get_column_letter(col)
            if ws.column_dimensions[col_letter].width:
                sheet_info['column_widths'][col_letter] = ws.column_dimensions[col_letter].width
        
        # Get row heights for first 20 rows
        for row in range(1, min(ws.max_row + 1, 20)):
            if ws.row_dimensions[row].height:
                sheet_info['row_heights'][row] = ws.row_dimensions[row].height
        
        # Sample first few cells for formatting
        for row in range(1, min(ws.max_row + 1, 6)):
            for col in range(1, min(ws.max_column + 1, 10)):
                cell = ws.cell(row, col)
                if cell.value:
                    cell_ref = f"{get_column_letter(col)}{row}"
                    sheet_info['sample_cells'][cell_ref] = {
                        'value': str(cell.value)[:50],  # Truncate long values
                        'font_bold': cell.font.bold if cell.font else None,
                        'font_size': cell.font.size if cell.font else None,
                        'alignment_horizontal': cell.alignment.horizontal if cell.alignment else None,
                        'alignment_vertical': cell.alignment.vertical if cell.alignment else None,
                        'alignment_wrap_text': cell.alignment.wrap_text if cell.alignment else None,
                        'fill_color': cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None,
                        'border': bool(cell.border and (cell.border.left or cell.border.right or cell.border.top or cell.border.bottom)),
                    }
        
        analysis['sheets'][sheet_name] = sheet_info
    
    return analysis

def compare_workbooks(original_file, generated_file):
    """Compare two workbooks and highlight differences"""
    print("=" * 80)
    print("EXCEL FORMATTING COMPARISON")
    print("=" * 80)
    print()
    
    original = analyze_workbook(original_file)
    generated = analyze_workbook(generated_file)
    
    print(f"📄 ORIGINAL:  {original_file}")
    print(f"📄 GENERATED: {generated_file}")
    print()
    
    # Compare sheet names
    print("📋 SHEETS")
    print("-" * 80)
    print(f"Original sheets:  {list(original['sheets'].keys())}")
    print(f"Generated sheets: {list(generated['sheets'].keys())}")
    print()
    
    # Compare each sheet (match by position if names differ)
    orig_sheet_names = list(original['sheets'].keys())
    gen_sheet_names = list(generated['sheets'].keys())
    
    for i, orig_sheet_name in enumerate(orig_sheet_names):
        if i >= len(gen_sheet_names):
            print(f"⚠️  Sheet {i+1} '{orig_sheet_name}' missing in generated file!")
            continue
        
        gen_sheet_name = gen_sheet_names[i]
        orig_sheet = original['sheets'][orig_sheet_name]
        gen_sheet = generated['sheets'][gen_sheet_name]
        
        sheet_name = f"{orig_sheet_name} vs {gen_sheet_name}"
        
        print(f"📊 SHEET: {sheet_name}")
        print("-" * 80)
        
        # Dimensions
        print(f"Dimensions:")
        print(f"  Original:  {orig_sheet['dimensions']['max_row']} rows × {orig_sheet['dimensions']['max_column']} cols")
        print(f"  Generated: {gen_sheet['dimensions']['max_row']} rows × {gen_sheet['dimensions']['max_column']} cols")
        print()
        
        # Frozen panes
        print(f"Frozen Panes:")
        print(f"  Original:  {orig_sheet['frozen_panes']}")
        print(f"  Generated: {gen_sheet['frozen_panes']}")
        if orig_sheet['frozen_panes'] != gen_sheet['frozen_panes']:
            print(f"  ❌ DIFFERENCE!")
        print()
        
        # Column widths
        print(f"Column Widths (first 10 columns):")
        all_cols = set(list(orig_sheet['column_widths'].keys()) + list(gen_sheet['column_widths'].keys()))
        for col in sorted(all_cols, key=lambda x: openpyxl.utils.column_index_from_string(x))[:10]:
            orig_width = orig_sheet['column_widths'].get(col, 'default')
            gen_width = gen_sheet['column_widths'].get(col, 'default')
            match = "✅" if orig_width == gen_width else "❌"
            print(f"  {col}: {orig_width:>8} | {gen_width:>8} {match}")
        print()
        
        # Merged cells
        print(f"Merged Cells:")
        print(f"  Original:  {len(orig_sheet['merged_cells'])} merged regions")
        print(f"  Generated: {len(gen_sheet['merged_cells'])} merged regions")
        if orig_sheet['merged_cells']:
            print(f"  Original regions: {orig_sheet['merged_cells'][:5]}")
        if gen_sheet['merged_cells']:
            print(f"  Generated regions: {gen_sheet['merged_cells'][:5]}")
        print()
        
        # Sample cell formatting
        print(f"Sample Cell Formatting (first 5 cells with values):")
        sample_cells = list(orig_sheet['sample_cells'].keys())[:5]
        for cell_ref in sample_cells:
            print(f"  {cell_ref}: {orig_sheet['sample_cells'][cell_ref]['value'][:30]}")
            orig_cell = orig_sheet['sample_cells'][cell_ref]
            gen_cell = gen_sheet['sample_cells'].get(cell_ref, {})
            
            if orig_cell.get('font_bold'):
                print(f"    Bold: {orig_cell['font_bold']} | {gen_cell.get('font_bold', 'N/A')} {'✅' if orig_cell['font_bold'] == gen_cell.get('font_bold') else '❌'}")
            if orig_cell.get('alignment_wrap_text'):
                print(f"    Wrap text: {orig_cell['alignment_wrap_text']} | {gen_cell.get('alignment_wrap_text', 'N/A')} {'✅' if orig_cell['alignment_wrap_text'] == gen_cell.get('alignment_wrap_text') else '❌'}")
            if orig_cell.get('fill_color'):
                print(f"    Fill: {orig_cell['fill_color']} | {gen_cell.get('fill_color', 'N/A')} {'✅' if orig_cell['fill_color'] == gen_cell.get('fill_color') else '❌'}")
        
        print()
        print()

if __name__ == '__main__':
    original = '1 Hour Ethogram Section(2).xlsx'
    generated = 'WBS-Ethogram-TestUser-2025-11-29-63107895.xlsx'
    
    try:
        compare_workbooks(original, generated)
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
