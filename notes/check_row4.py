#!/usr/bin/env python3
"""Quick check - does the original have bold headers in row 4?"""
import openpyxl

wb = openpyxl.load_workbook('1 Hour Ethogram Section(2).xlsx')
ws = wb[wb.sheetnames[0]]

print("Checking row 4 (time slot headers):")
for col in range(2, 14):  # Columns B-M
    cell = ws.cell(4, col)
    is_bold = cell.font.bold if cell.font else None
    print(f"  Column {openpyxl.utils.get_column_letter(col)}: value='{cell.value}', bold={is_bold}")
