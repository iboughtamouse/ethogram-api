#!/usr/bin/env python3
"""
Analyze Excel file structure to compare original vs generated format
"""

import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border

def analyze_excel(filepath):
    print(f"\n{'='*80}")
    print(f"Analyzing: {filepath}")
    print('='*80)

    wb = openpyxl.load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"Dimensions: {ws.dimensions}")
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

        # Show first 30 rows
        print("\nFirst 30 rows:")
        for row_idx in range(1, min(31, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(15, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                value = str(cell.value) if cell.value else ""

                # Show cell styling info for non-empty cells
                if value:
                    style_info = []
                    if cell.font and cell.font.bold:
                        style_info.append("BOLD")
                    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                        style_info.append(f"BG:{cell.fill.fgColor.rgb}")
                    if cell.alignment and cell.alignment.wrap_text:
                        style_info.append("WRAP")

                    if style_info:
                        value = f"{value} [{','.join(style_info)}]"

                row_data.append(value[:40] if value else "")

            print(f"Row {row_idx:2d}: {' | '.join(row_data)}")

        # Show column widths
        print("\nColumn widths:")
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            width = ws.column_dimensions[col_letter].width
            if width:
                print(f"  {col_letter}: {width}")

        # Show merged cells
        if ws.merged_cells:
            print("\nMerged cells:")
            for merged_range in ws.merged_cells.ranges:
                print(f"  {merged_range}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python analyze_excel.py <file1.xlsx> [file2.xlsx]")
        sys.exit(1)

    for filepath in sys.argv[1:]:
        try:
            analyze_excel(filepath)
        except Exception as e:
            print(f"\nError analyzing {filepath}: {e}")
